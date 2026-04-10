import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border

def generate_packing_list(data: dict, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List"
    
    # ---------------- FONTS & FILLS ----------------
    font_title = Font(name='Arial', size=16, bold=True)
    font_header = Font(name='Arial', size=10, bold=True)
    font_normal = Font(name='Arial', size=10)
    
    fill_blue = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    header_info = data.get('header_information', {})
    line_items = data.get('line_items', [])

    # ROW 1: Title
    ws.merge_cells('A1:Q1')
    ws['A1'] = "PACKING LIST"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center

    # ROW 2-15: Header section (Left / Right split roughly)
    # Left: A to H, Right: I to Q
    ws.merge_cells('A2:H7')
    ws['A2'] = f"Exporter Name & Address:\\n{header_info.get('exporter_name_address', 'N/A')}\\nGST: {header_info.get('exporter_gst', 'N/A')}\\nIEC: {header_info.get('exporter_iec', 'N/A')}"
    ws['A2'].alignment = align_top_left
    ws['A2'].font = font_normal
    
    ws.merge_cells('I2:Q3')
    ws['I2'] = f"Invoice No. & Date: {header_info.get('invoice_number_date', 'N/A')}"
    ws['I2'].alignment = align_left
    ws['I2'].font = font_normal

    ws.merge_cells('I4:Q5')
    ws['I4'] = f"Buyer's PO No. & Date: {header_info.get('buyer_po_number_date', 'N/A')}"
    ws['I4'].alignment = align_left
    ws['I4'].font = font_normal

    ws.merge_cells('A8:H13')
    ws['A8'] = f"Consignee Name & Address:\\n{header_info.get('consignee_name_address', 'N/A')}"
    ws['A8'].alignment = align_top_left
    ws['A8'].font = font_normal

    ws.merge_cells('A14:Q15')
    ws['A14'] = f"Buyer Name & Address (if other than Consignee):\\n{header_info.get('buyer_name_address', 'N/A')}"
    ws['A14'].alignment = align_left
    ws['A14'].font = font_normal

    # Apply blue fill to header section roughly
    for row in ws.iter_rows(min_row=2, max_row=15, min_col=1, max_col=17):
        for cell in row:
            cell.fill = fill_blue

    # ROW 16-18: Shipping Details
    ws.merge_cells('A16:D16')
    ws['A16'] = "Pre-carriage by:"
    ws['A16'].font = font_header
    
    ws.merge_cells('E16:H16')
    ws['E16'] = "Place of Receipt:"
    ws['E16'].font = font_header
    
    ws.merge_cells('I16:M16')
    ws['I16'] = f"Pre-carriage: N/A"
    ws['I16'].font = font_normal

    ws.merge_cells('N16:Q16')
    ws['N16'] = f"Receipt: N/A"
    ws['N16'].font = font_normal

    ws.merge_cells('A17:D17')
    ws['A17'] = f"Port of Loading:\\n{header_info.get('port_of_loading', 'N/A')}"
    ws.merge_cells('E17:H17')
    ws['E17'] = f"Port of Discharge:\\n{header_info.get('port_of_discharge', 'N/A')}"
    ws.merge_cells('I17:M17')
    ws['I17'] = f"Country of Origin:\\n{header_info.get('country_of_origin', 'N/A')}"
    ws.merge_cells('N17:Q17')
    ws['N17'] = f"Country of Destination:\\n{header_info.get('country_of_destination', 'N/A')}"

    ws.merge_cells('A18:Q18')
    ws['A18'] = f"Terms of Delivery: {header_info.get('terms_of_delivery', 'N/A')} | Vessel/Flight: {header_info.get('vessel_flight_details', 'N/A')} | Marks: {header_info.get('shipping_marks', 'N/A')}"
    ws['A18'].font = font_normal

    for row in ws.iter_rows(min_row=16, max_row=18, min_col=1, max_col=17):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = align_left

    # TABLE HEADERS
    headers = [
        "Sr. No", "PO Number", "Style No", "Description", "Color", 
        "Size XS", "Size S", "Size M", "Size L", "Size XL", "Size XXL",
        "Total Qty", "No of Cartons", "Carton Size (LxWxH cm)", "Net Wt (Kgs)", "Gross Wt (Kgs)", "CBM"
    ]
    
    start_row = 20
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_thin

    # DATA ROWS
    current_row = start_row + 1
    total_qty = 0
    total_cartons = 0
    total_net = 0.0
    total_gross = 0.0
    total_cbm = 0.0

    for idx, item in enumerate(line_items, 1):
        sizes = item.get('sizes', {})
        qty = item.get('total_quantity', 0)
        
        # Mock calculation since PO might not have carton dimensions, assume something standard if missing or put N/A
        # Default carton calc just as an example logic
        ctn = max(1, qty // 50) if qty > 0 else 0
        l, w, h = 60, 40, 40
        net = qty * 0.2
        gross = net + ctn * 1.5
        cbm = (l * w * h) / 1000000 * ctn

        row_data = [
            idx,
            header_info.get('buyer_po_number_date', 'N/A'),
            item.get('style_number', 'N/A'),
            item.get('product_description', 'N/A'),
            item.get('color', 'N/A'),
            sizes.get('XS', 0), sizes.get('S', 0), sizes.get('M', 0),
            sizes.get('L', 0), sizes.get('XL', 0), sizes.get('XXL', 0),
            qty, ctn, f"{l}x{w}x{h}", round(net, 2), round(gross, 2), round(cbm, 4)
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_normal
            cell.alignment = align_center
            cell.border = border_thin

        total_qty += qty
        total_cartons += ctn
        total_net += net
        total_gross += gross
        total_cbm += cbm
        
        current_row += 1

    # TOTAL ROW
    for col_idx in range(1, 18):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = fill_yellow
        cell.font = font_header
        cell.border = border_thin
        cell.alignment = align_center

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
    ws.cell(row=current_row, column=1, value="TOTAL").alignment = align_center
    
    ws.cell(row=current_row, column=12, value=total_qty)
    ws.cell(row=current_row, column=13, value=total_cartons)
    ws.cell(row=current_row, column=15, value=round(total_net, 2))
    ws.cell(row=current_row, column=16, value=round(total_gross, 2))
    ws.cell(row=current_row, column=17, value=round(total_cbm, 4))
    
    current_row += 3
    # FOOTER DECLARATION
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=17)
    decl_cell = ws.cell(row=current_row, column=1, value="We hereby certify that the goods described above are of Indian Origin")
    decl_cell.font = font_normal
    decl_cell.alignment = align_center

    current_row += 4
    ws.merge_cells(start_row=current_row, start_column=13, end_row=current_row, end_column=17)
    sig_cell = ws.cell(row=current_row, column=13, value="For " + header_info.get('exporter_name_address', '').split('\\n')[0])
    sig_cell.font = font_header
    sig_cell.alignment = Alignment(horizontal="right")
    
    current_row += 2
    ws.merge_cells(start_row=current_row, start_column=14, end_row=current_row, end_column=17)
    auth_cell = ws.cell(row=current_row, column=14, value="Authorized Signatory")
    auth_cell.font = font_normal
    auth_cell.alignment = Alignment(horizontal="center")

    # Layout adjustments
    for col in range(1, 18):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    ws.column_dimensions['D'].width = 25 # Description
    ws.column_dimensions['A'].width = 8
    
    # Page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

    wb.save(output_path)
